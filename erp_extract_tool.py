#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ERP凭证字段提取工具 v3.0
批量从ERP导出的Excel文件中提取指定字段，合并导出到一个新Excel文件。
支持拖拽文件、进度条、完成提示、字段多选、偏好记忆。
"""

import os
import sys
import json
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

# 轻型库立即导入（GUI 不阻塞）
from tkinterdnd2 import TkinterDnD, DND_FILES

# ── 默认目标字段列表 ─────────────────────────────────────────
DEFAULT_TARGET_COLUMNS = [
    "凭证抬头文本",
    "文本",
    "凭证货币价值",
    "借/贷标识",
    "总账科目",
    "总账科目：长文本",
]

# ── 允许拖入的文件扩展名 ────────────────────────────────────
ALLOWED_EXTS = {".xlsx", ".xls"}

# ── 偏好文件路径 ────────────────────────────────────────────
PREFS_DIR = os.path.expanduser("~/Library/Application Support/ERP-Extract-Tool")
PREFS_PATH = os.path.join(PREFS_DIR, "prefs.json")


def load_prefs():
    """加载偏好文件，返回 selected_columns 列表或 None。"""
    try:
        if os.path.exists(PREFS_PATH):
            with open(PREFS_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            cols = data.get("selected_columns", [])
            if isinstance(cols, list) and cols:
                return cols
    except Exception:
        pass
    return None


def save_prefs(selected_columns):
    """将勾选的字段列表写入偏好文件。"""
    try:
        os.makedirs(PREFS_DIR, exist_ok=True)
        with open(PREFS_PATH, "w", encoding="utf-8") as f:
            json.dump({"selected_columns": selected_columns}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def find_column_mapping(headers, targets):
    """在表头列表中查找目标字段的列索引。先精确匹配，再子串匹配。"""
    mapping = {}
    cleaned = {i: str(h).strip() if h is not None else "" for i, h in enumerate(headers)}
    for target in targets:
        found = False
        for idx, name in cleaned.items():
            if name == target:
                mapping[target] = idx
                found = True
                break
        if not found:
            for idx, name in cleaned.items():
                if target in name:
                    mapping[target] = idx
                    found = True
                    break
    return mapping


def parse_drop_data(data):
    """解析拖拽事件传来的文件路径列表。macOS 下 tkinterdnd2 返回带花括号的路径。"""
    paths = []
    raw = data.strip()
    if not raw:
        return paths
    brace_depth = 0
    current = ""
    for ch in raw:
        if ch == "{":
            brace_depth += 1
            if brace_depth == 1:
                continue
        elif ch == "}":
            brace_depth -= 1
            if brace_depth == 0:
                if current:
                    paths.append(current)
                current = ""
                continue
        if brace_depth > 0:
            current += ch
    if not paths:
        for item in raw.split():
            item = item.strip()
            if item:
                paths.append(item)
    return paths


# ═══════════════════ GUI 应用 ═══════════════════

class ErpExtractApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ERP凭证字段提取工具")
        self.root.geometry("950x640")
        self.root.resizable(True, True)

        self.files = []                  # [(显示名, 完整路径), ...]
        self.export_dir = os.path.expanduser("~/Desktop")
        self._heavy_loaded = False
        self._scan_after_id = None       # 延迟扫描的 after ID
        self._scanning = False           # 是否正在扫描表头
        self._intersection_headers = []  # 交集表头列表
        self.field_vars = {}             # {字段名: tk.BooleanVar}
        self._checkbutton_widgets = []   # 当前显示的 Checkbutton 控件

        # 加载偏好
        self._saved_prefs = load_prefs()

        self._build_ui()
        self.root.after(50, self._lazy_load_heavy_libs)

    # ── 延迟加载重型库 ──────────────────────────────────────

    def _lazy_load_heavy_libs(self):
        def load():
            global pd, openpyxl
            import pandas as _pd
            import openpyxl as _oxl
            pd = _pd
            openpyxl = _oxl
            self._heavy_loaded = True
            self.root.after(0, lambda: self._append_log("[引擎就绪] pandas + openpyxl 已加载"))
            # 引擎就绪后，如果有文件待扫描则触发
            self.root.after(0, self._on_engine_ready)
        threading.Thread(target=load, daemon=True).start()

    def _on_engine_ready(self):
        if self.files:
            self._schedule_header_scan()

    # ── 构建界面 ────────────────────────────────────────────

    def _build_ui(self):
        # ── 主内容区：左侧文件列表 + 右侧字段选择 ──
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 0))

        # 左侧：文件列表区 (40%)
        left_frame = ttk.LabelFrame(main_frame, text="待处理文件列表（可拖拽 .xlsx/.xls 到此处）", padding=5)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        list_frame = ttk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        self.listbox = tk.Listbox(
            list_frame,
            selectmode=tk.EXTENDED,
            yscrollcommand=scrollbar.set,
            font=("Monaco", 11),
        )
        scrollbar.config(command=self.listbox.yview)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 注册拖拽目标
        self.listbox.drop_target_register(DND_FILES)
        self.listbox.dnd_bind("<<Drop>>", self._on_drop)

        # 文件操作按钮
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(btn_frame, text="添加文件", command=self.add_files).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_frame, text="移除选中", command=self.remove_selected).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="清空列表", command=self.clear_all).pack(side=tk.LEFT, padx=(5, 0))

        # 右侧：字段选择区 (60%)
        right_frame = ttk.LabelFrame(main_frame, text="字段选择", padding=5)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))

        # Canvas + Scrollbar 实现可滚动字段区域
        field_canvas_frame = ttk.Frame(right_frame)
        field_canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.field_canvas = tk.Canvas(field_canvas_frame, height=200, highlightthickness=0)
        self.field_scrollbar = ttk.Scrollbar(field_canvas_frame, orient=tk.VERTICAL, command=self.field_canvas.yview)
        self.field_canvas.configure(yscrollcommand=self.field_scrollbar.set)

        self.field_inner = ttk.Frame(self.field_canvas)
        self.field_inner.bind("<Configure>", lambda e: self.field_canvas.configure(
            scrollregion=self.field_canvas.bbox("all")
        ))
        self.field_canvas.create_window((0, 0), window=self.field_inner, anchor="nw", tags="inner")
        # 同步 inner frame 宽度
        self.field_canvas.bind("<Configure>", lambda e: self.field_canvas.itemconfig(
            "inner", width=e.width
        ))

        self.field_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.field_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 占位提示
        self._field_placeholder = ttk.Label(self.field_inner, text="请先添加文件以显示可用字段", foreground="gray")
        self._field_placeholder.pack(pady=10)

        # 快捷按钮
        field_btn_frame = ttk.Frame(right_frame)
        field_btn_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(field_btn_frame, text="全选", command=self._select_all_fields).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(field_btn_frame, text="反选", command=self._invert_fields).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(field_btn_frame, text="重置默认（6字段）", command=self._reset_default_fields).pack(side=tk.LEFT)

        # ── 底部区域：导出目录 + 进度条 + 按钮 + 日志 ──
        bottom_frame = ttk.Frame(self.root)
        bottom_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 导出目录
        dir_frame = ttk.LabelFrame(bottom_frame, text="导出设置", padding=5)
        dir_frame.pack(fill=tk.X, pady=(0, 5))

        dir_row = ttk.Frame(dir_frame)
        dir_row.pack(fill=tk.X)
        self.export_label = tk.Label(
            dir_row, text=self.export_dir, anchor="w", relief="sunken",
            bg="white", fg="gray", font=("Monaco", 10),
        )
        self.export_label.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=2)
        ttk.Button(dir_row, text="选择导出目录", command=self.choose_export_dir).pack(side=tk.RIGHT, padx=(5, 0))

        # 进度条
        progress_frame = ttk.Frame(bottom_frame)
        progress_frame.pack(fill=tk.X, pady=(0, 5))

        self.progress_var = tk.DoubleVar(value=0)
        self.progress = ttk.Progressbar(
            progress_frame, variable=self.progress_var, mode="determinate", maximum=100,
        )
        self.progress.pack(fill=tk.X)

        self.progress_label = ttk.Label(progress_frame, text="就绪", anchor="w", font=("Monaco", 9))
        self.progress_label.pack(fill=tk.X, pady=(2, 0))

        # 开始提取按钮
        self.btn_start = ttk.Button(bottom_frame, text="开始提取", command=self.start_extraction)
        self.btn_start.pack(pady=(0, 5))

        # 运行日志
        log_frame = ttk.LabelFrame(bottom_frame, text="运行日志", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.status_text = tk.Text(
            log_frame, height=5, wrap=tk.WORD, font=("Monaco", 10), state=tk.DISABLED,
        )
        self.status_text.pack(fill=tk.BOTH, expand=True)

        # 绑定鼠标滚轮
        self.field_canvas.bind("<Enter>", self._bind_mousewheel)
        self.field_canvas.bind("<Leave>", self._unbind_mousewheel)

    def _bind_mousewheel(self, event):
        self.field_canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _unbind_mousewheel(self, event):
        self.field_canvas.unbind_all("<MouseWheel>")

    def _on_mousewheel(self, event):
        self.field_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ── 拖拽回调 ────────────────────────────────────────────

    def _on_drop(self, event):
        paths = parse_drop_data(event.data)
        added = 0
        for p in paths:
            ext = os.path.splitext(p)[1].lower()
            if ext not in ALLOWED_EXTS:
                continue
            if p not in [fp for _, fp in self.files]:
                self.files.append((os.path.basename(p), p))
                added += 1
        if added:
            self._refresh_list()
            self._append_log(f"[拖入] {added} 个文件")
            self._schedule_header_scan()

    # ── 按钮回调 ────────────────────────────────────────────

    def add_files(self):
        paths = filedialog.askopenfilenames(
            title="选择ERP凭证Excel文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
        )
        added = 0
        for p in paths:
            if p not in [fp for _, fp in self.files]:
                self.files.append((os.path.basename(p), p))
                added += 1
        if added:
            self._refresh_list()
            self._schedule_header_scan()

    def remove_selected(self):
        selected = self.listbox.curselection()
        if not selected:
            return
        for i in reversed(selected):
            del self.files[i]
        self._refresh_list()
        self._schedule_header_scan()

    def clear_all(self):
        if not self.files:
            return
        self.files.clear()
        self._refresh_list()
        self._intersection_headers = []
        self._rebuild_checkbuttons()

    def choose_export_dir(self):
        d = filedialog.askdirectory(title="选择导出目录")
        if d:
            self.export_dir = d
            self.export_label.config(text=d)

    # ── 表头扫描 ────────────────────────────────────────────

    def _schedule_header_scan(self):
        """延迟触发表头扫描（防抖 300ms）。"""
        if self._scan_after_id is not None:
            self.root.after_cancel(self._scan_after_id)
        self._scan_after_id = self.root.after(300, self._do_header_scan)

    def _do_header_scan(self):
        """在后台线程扫描所有文件的表头，计算交集。"""
        if not self._heavy_loaded:
            # 引擎未就绪，等就绪后自动触发
            self._append_log("[等待引擎就绪后扫描表头...]")
            return
        if self._scanning:
            return
        if not self.files:
            self._intersection_headers = []
            self._rebuild_checkbuttons()
            return

        self._scanning = True
        self._append_log("[扫描表头中...]")

        def scan():
            file_paths = [fp for _, fp in self.files]
            all_header_sets = []
            first_file_header_order = []  # 第一个文件的表头原始顺序
            first = True
            for fpath in file_paths:
                try:
                    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                    for sname in wb.sheetnames:
                        try:
                            df = pd.read_excel(
                                fpath, sheet_name=sname, header=0, dtype=str,
                                engine="openpyxl", nrows=0,
                            )
                            headers = [str(h).strip() if h is not None else "" for h in df.columns.tolist()]
                            if headers:
                                all_header_sets.append(set(headers))
                                if first:
                                    first_file_header_order = headers
                            break  # 只看第一个 sheet 的表头
                        except Exception:
                            continue
                    wb.close()
                except Exception:
                    continue
                first = False

            if all_header_sets:
                intersection = all_header_sets[0]
                for hs in all_header_sets[1:]:
                    intersection = intersection & hs
                # 按第一个文件的表头原始顺序排列; 不在第一个文件中的(理论上不会)放在末尾
                ordered = [h for h in first_file_header_order if h in intersection]
                extras = [h for h in intersection if h not in ordered]
                result = ordered + sorted(extras)
            else:
                result = []

            self.root.after(0, lambda: self._on_headers_scanned(result))

        threading.Thread(target=scan, daemon=True).start()

    def _on_headers_scanned(self, headers):
        """表头扫描完成回调（主线程）。"""
        self._scanning = False
        self._intersection_headers = headers
        if headers:
            self._append_log(f"[表头扫描完成] 交集字段共 {len(headers)} 个")
        else:
            self._append_log("[表头扫描完成] 未找到交集字段")
        self._rebuild_checkbuttons()

    # ── 字段选择区域 ────────────────────────────────────────

    def _rebuild_checkbuttons(self):
        """根据交集表头重建 Checkbutton 列表。"""
        # 清除旧控件
        for w in self._checkbutton_widgets:
            w.destroy()
        self._checkbutton_widgets.clear()
        self.field_vars.clear()

        # 隐藏或显示占位提示
        if self._field_placeholder:
            self._field_placeholder.destroy()
            self._field_placeholder = None

        if not self._intersection_headers:
            self._field_placeholder = ttk.Label(self.field_inner, text="请先添加文件以显示可用字段", foreground="gray")
            self._field_placeholder.pack(pady=10)
            return

        # 确定默认勾选：偏好 > 默认 6 字段
        if self._saved_prefs is not None:
            default_set = set(self._saved_prefs) & set(self._intersection_headers)
        else:
            default_set = set(DEFAULT_TARGET_COLUMNS) & set(self._intersection_headers)

        # 如果偏好/默认与交集无交集，则全部不勾选
        for field in self._intersection_headers:
            var = tk.BooleanVar(value=(field in default_set))
            self.field_vars[field] = var
            cb = ttk.Checkbutton(self.field_inner, text=field, variable=var)
            cb.pack(anchor="w", padx=5, pady=1)
            self._checkbutton_widgets.append(cb)

        # 更新 scrollregion
        self.field_inner.update_idletasks()
        self.field_canvas.configure(scrollregion=self.field_canvas.bbox("all"))

        self._append_log(f"[字段列表已更新] 共 {len(self._intersection_headers)} 个交集字段，"
                         f"默认勾选 {sum(1 for v in self.field_vars.values() if v.get())} 个")

    def _get_selected_columns(self):
        """返回当前勾选的字段列表。"""
        return [name for name, var in self.field_vars.items() if var.get()]

    def _select_all_fields(self):
        for var in self.field_vars.values():
            var.set(True)

    def _invert_fields(self):
        for var in self.field_vars.values():
            var.set(not var.get())

    def _reset_default_fields(self):
        default_set = set(DEFAULT_TARGET_COLUMNS)
        for name, var in self.field_vars.items():
            var.set(name in default_set)

    # ── 提取逻辑 ────────────────────────────────────────────

    def start_extraction(self):
        if not self.files:
            messagebox.showwarning("提示", "请先添加要处理的Excel文件。")
            return
        if not self._heavy_loaded:
            messagebox.showinfo("请稍候", "处理引擎正在加载中，请稍后再试。")
            return

        selected_cols = self._get_selected_columns()
        if not selected_cols:
            messagebox.showwarning("提示", "请至少勾选一个要提取的字段。")
            return

        # 保存偏好
        save_prefs(selected_cols)

        self.btn_start.config(state=tk.DISABLED)
        self.status_text.config(state=tk.NORMAL)
        self.status_text.delete("1.0", tk.END)

        self.progress_var.set(0)
        self.progress_label.config(text="正在提取...")
        self.progress.configure(style="TProgressbar")

        total = len(self.files)
        self.progress["maximum"] = total

        def log(msg):
            self.root.after(0, lambda: self._append_log(msg))

        def progress_update(idx, fname):
            self.root.after(0, lambda: self._update_progress(idx, fname))

        def done():
            self.root.after(0, lambda: self.btn_start.config(state=tk.NORMAL))

        def worker():
            log("开始提取...")
            out, records = self._process_files(progress_update, log, selected_cols)
            if out:
                log(f"\n导出文件: {out}")
                self.root.after(0, lambda: self._on_complete(total, records, out))
            else:
                self.root.after(0, lambda: self._reset_progress())
            done()

        threading.Thread(target=worker, daemon=True).start()

    def _process_files(self, progress_callback, log_callback, target_columns):
        """核心提取逻辑。返回 (输出路径, 记录数) 或 (None, 0)。"""
        all_rows = []
        file_paths = [fp for _, fp in self.files]

        for fi, file_path in enumerate(file_paths, 1):
            fname = os.path.basename(file_path)
            progress_callback(fi, fname)

            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                wb.close()

                for sname in sheet_names:
                    try:
                        df = pd.read_excel(
                            file_path, sheet_name=sname, header=0, dtype=str,
                            engine="openpyxl",
                        )
                    except Exception:
                        log_callback(f"  跳过 sheet「{sname}」（读取失败）")
                        continue

                    if df.empty:
                        continue

                    headers = df.columns.tolist()
                    col_map = find_column_mapping(headers, target_columns)
                    if not col_map:
                        continue

                    for _, row in df.iterrows():
                        row_data = {"来源文件": fname}
                        for target in target_columns:
                            idx = col_map.get(target)
                            if idx is not None and idx < len(row):
                                val = row.iloc[idx]
                                if pd.isna(val):
                                    val = ""
                                else:
                                    val = str(val).strip()
                            else:
                                val = ""
                            row_data[target] = val
                        all_rows.append(row_data)

            except Exception as e:
                log_callback(f"  处理文件失败: {fname} - {e}")

        if not all_rows:
            log_callback("未提取到任何数据，请检查文件格式。")
            return None, 0

        out_columns = ["来源文件"] + target_columns
        result_df = pd.DataFrame(all_rows, columns=out_columns)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(self.export_dir, f"汇总_{ts}.xlsx")

        try:
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                result_df.to_excel(writer, index=False, sheet_name="汇总")
            log_callback(f"完成！共提取 {len(all_rows)} 条记录，导出至:\n  {out_path}")
            return out_path, len(all_rows)
        except Exception as e:
            log_callback(f"写入文件失败: {e}")
            return None, 0

    def _update_progress(self, idx, fname):
        self.progress_var.set(idx)
        self.progress_label.config(text=f"[{idx}/{int(self.progress['maximum'])}] {fname}")

    def _on_complete(self, file_count, record_count, out_path):
        self.progress_var.set(self.progress["maximum"])
        self.progress_label.config(text="提取完成")

        style = ttk.Style()
        style_name = "green.Horizontal.TProgressbar"
        style.configure(style_name, troughcolor="white", background="#4CAF50")
        self.progress.configure(style=style_name)

        os.system(f'open "{self.export_dir}"')

        messagebox.showinfo(
            "提取完成",
            f"处理了 {file_count} 个文件\n共提取 {record_count} 条记录\n\n导出路径:\n{out_path}",
        )
        self._append_log(f"\n--- 完成：{file_count} 文件, {record_count} 条记录 ---")

    def _reset_progress(self):
        self.progress_var.set(0)
        self.progress_label.config(text="就绪")
        self.progress.configure(style="TProgressbar")

    # ── 日志/列表刷新 ───────────────────────────────────────

    def _append_log(self, msg):
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, msg + "\n")
        self.status_text.see(tk.END)
        self.status_text.config(state=tk.DISABLED)

    def _refresh_list(self):
        self.listbox.delete(0, tk.END)
        for name, _ in self.files:
            self.listbox.insert(tk.END, name)


def main():
    root = TkinterDnD.Tk()
    app = ErpExtractApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
